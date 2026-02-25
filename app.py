"""
Project RAG Pro — OpenAI Assistants + File Search Edition
=========================================================
Multi-project RAG with dual document pools (Main + Clarification).
Uses OpenAI Assistants API with File Search for chunking, embedding,
storage, and retrieval — all handled by OpenAI's infrastructure.

Dependencies: flask, openai, openpyxl
"""

from __future__ import annotations

import json
import logging
import os
import shutil
import tempfile
import threading
import time
from pathlib import Path
from typing import Any

os.environ.setdefault("KMP_DUPLICATE_LIB_OK", "TRUE")

from flask import Flask, jsonify, render_template, request, session, redirect, url_for
from openai import OpenAI

# ── Configuration ────────────────────────────────────────────────────────────

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
CHAT_MODEL = "gpt-4.1-mini"
APP_PASSWORD = os.getenv("APP_PASSWORD", "")
SECRET_KEY = os.getenv("SECRET_KEY", "rag-pro-secret-key-change-me")
BASE_DIR = Path(os.getenv("PROJECTS_DIR", Path(__file__).resolve().parent / "projects"))
POOL_TYPES = ("main", "clarification")

# ── Logging ──────────────────────────────────────────────────────────────────

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(message)s")
log = logging.getLogger("rag")

# ── Validate OpenAI ──────────────────────────────────────────────────────────

if not OPENAI_API_KEY:
    print("\n  [ERROR] OPENAI_API_KEY environment variable is not set.\n")
    raise SystemExit(1)

client = OpenAI()

print(f"""
╔══════════════════════════════════════════════╗
║   Project RAG Pro — Assistants Edition       ║
╠══════════════════════════════════════════════╣
║  Chat model:  {CHAT_MODEL:<30s}  ║
║  Backend:     OpenAI Assistants + File Search║
║  Projects in: {str(BASE_DIR):<30s}  ║
╚══════════════════════════════════════════════╝
""")

# ── System Prompt ────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """You are an Expert Bid Review Engineer with 25+ years of experience in EPC 
(Engineering, Procurement, and Construction) projects across Oil & Gas, Petrochemical, 
Power, and Infrastructure sectors.

YOUR MISSION: Extract EVERY piece of information from the project documents with ZERO data loss.
These are client bid documents — missing any detail could mean missing a critical requirement.

ANALYSIS METHODOLOGY — Follow these 5 steps for EVERY question:
1. SCAN: Read ALL retrieved document snippets thoroughly, treating each as potentially critical.
2. EXTRACT: Pull out every relevant data point — specifications, dimensions, quantities, dates, 
   codes, standards, scope items, commercial terms, schedules, and referenced documents.
3. CONSOLIDATE: Cross-reference information from multiple snippets to build a complete picture.
4. STRUCTURE: Organize your answer with clear headings and numbered points for complex responses.
5. VERIFY: Re-scan all snippets one final time before submitting to ensure nothing was missed.

MANDATORY RULES:
- EXHAUSTIVE EXTRACTION: Always provide ALL details found, never summarize away specifics.
- CITE EVERYTHING: Every data point must reference the source file.
- FLAG DISCREPANCIES: If two documents give different values, report BOTH and flag the contradiction.
- FLAG MISSING DATA: If only partial information is found, explicitly list what is missing.
- COMPLETENESS OVER BREVITY: It is better to include too much than to miss a requirement.
- TECHNICAL PRECISION: Use exact values, units, and terminology from the documents.
- When you find relevant information, present it in structured tables where appropriate.
- If the documents do not contain the answer, say so clearly — do NOT fabricate information."""

# ── Helper: Excel to text conversion ─────────────────────────────────────────

def _read_excel_with_pandas(filepath: Path) -> dict[str, list[list[str]]]:
    """Read an Excel file using pandas — handles corrupt custom properties that crash openpyxl.
    
    Returns a dict of {sheet_name: [[row1_cells], [row2_cells], ...]}.
    """
    import pandas as pd
    sheets = {}
    try:
        # Read all sheets; header=None so we get raw rows including the header row
        xls = pd.ExcelFile(filepath, engine="openpyxl")
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name, header=None, dtype=str)
            df = df.fillna("")
            rows = df.values.tolist()
            sheets[sheet_name] = rows
        xls.close()
    except Exception as e:
        log.warning(f"pandas/openpyxl failed for {filepath.name}: {e}")
        # Fallback: patch openpyxl custom properties to bypass the bug
        try:
            sheets = _read_excel_with_patched_openpyxl(filepath)
        except Exception as e2:
            log.warning(f"Patched openpyxl also failed for {filepath.name}: {e2}")
            # Last resort: try calamine engine if available
            try:
                xls = pd.ExcelFile(filepath, engine="calamine")
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet_name, header=None, dtype=str)
                    df = df.fillna("")
                    sheets[sheet_name] = df.values.tolist()
                xls.close()
            except Exception as e3:
                log.error(f"All Excel readers failed for {filepath.name}: {e3}")
    return sheets


def _read_excel_with_patched_openpyxl(filepath: Path) -> dict[str, list[list[str]]]:
    """Read Excel by monkey-patching openpyxl to skip broken custom properties."""
    import openpyxl
    from openpyxl.packaging import custom
    
    # Save original
    _orig_from_tree = custom.CustomPropertyList.from_tree
    
    @classmethod
    def _safe_from_tree(cls, tree):
        """Patched version that skips properties with None names."""
        try:
            return _orig_from_tree(tree)
        except Exception:
            # Return empty property list if parsing fails
            return cls()
    
    # Apply patch
    custom.CustomPropertyList.from_tree = _safe_from_tree
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        sheets = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c).strip() if c is not None else "" for c in row])
            sheets[sheet_name] = rows
        wb.close()
        return sheets
    finally:
        # Restore original
        custom.CustomPropertyList.from_tree = _orig_from_tree


def excel_to_text(filepath: Path) -> str:
    """Convert an Excel file to a rich text representation for upload to OpenAI.
    
    Uses pandas as the primary reader (handles corrupt custom properties),
    with patched openpyxl and calamine as fallbacks.
    
    Produces a well-structured text document that preserves:
    - All sheet names
    - Column headers
    - Every row of data with labeled columns
    - The original filename for citation purposes
    """
    sheets = _read_excel_with_pandas(filepath)
    
    if not sheets:
        log.error(f"Could not read any sheets from {filepath.name}")
        return ""
    
    parts = []
    parts.append(f"DOCUMENT: {filepath.name}")
    parts.append(f"TYPE: Excel Spreadsheet")
    parts.append(f"SHEETS: {', '.join(sheets.keys())}")
    parts.append("=" * 80)
    
    total_rows = 0
    
    for sheet_name, rows in sheets.items():
        parts.append(f"\n{'=' * 80}")
        parts.append(f"SHEET: {sheet_name}")
        parts.append(f"{'=' * 80}\n")
        
        if not rows:
            parts.append("[Empty sheet]")
            continue
        
        # Find headers (first non-empty row)
        headers = None
        data_start = 0
        for i, row in enumerate(rows):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                headers = cells
                data_start = i + 1
                break
        
        if not headers:
            parts.append("[No data found in this sheet]")
            continue
        
        # Output as a table format
        parts.append("COLUMNS: " + " | ".join(headers))
        parts.append("-" * 80)
        
        sheet_rows = 0
        for row in rows[data_start:]:
            cells = [str(c).strip() if c is not None else "" for c in row]
            # Skip completely empty rows
            if not any(cells):
                continue
            
            # Format as labeled columns for better semantic search
            labeled_parts = []
            for j, cell in enumerate(cells):
                if j < len(headers) and headers[j]:
                    labeled_parts.append(f"{headers[j]}: {cell}")
                elif cell:
                    labeled_parts.append(cell)
            
            if labeled_parts:
                parts.append(" | ".join(labeled_parts))
                sheet_rows += 1
        
        parts.append(f"\n[Sheet '{sheet_name}': {sheet_rows} data rows]")
        total_rows += sheet_rows
    
    parts.append(f"\n{'=' * 80}")
    parts.append(f"END OF DOCUMENT: {filepath.name}")
    parts.append(f"TOTAL DATA ROWS: {total_rows}")
    parts.append(f"{'=' * 80}")
    
    result = "\n".join(parts)
    log.info(f"Excel conversion: {filepath.name} → {total_rows} rows, {len(result)} chars")
    return result


# ── Project metadata management ──────────────────────────────────────────────

def get_project_dir(project_id: str) -> Path:
    return BASE_DIR / project_id


def meta_path(project_dir: Path) -> Path:
    return project_dir / "metadata.json"


def load_meta(project_dir: Path) -> dict:
    mp = meta_path(project_dir)
    if mp.exists():
        return json.loads(mp.read_text())
    return {
        "description": "",
        "created": "",
        "pools": {
            "main": {"vector_store_id": "", "assistant_id": "", "files": {}},
            "clarification": {"vector_store_id": "", "assistant_id": "", "files": {}},
        },
    }


def save_meta(project_dir: Path, meta: dict):
    meta_path(project_dir).write_text(json.dumps(meta, indent=2))


# ── OpenAI resource management ───────────────────────────────────────────────

def ensure_vector_store(meta: dict, pool: str, project_name: str) -> str:
    """Ensure a vector store exists for the pool, create if needed."""
    pool_meta = meta["pools"][pool]
    vs_id = pool_meta.get("vector_store_id", "")
    if vs_id:
        try:
            client.vector_stores.retrieve(vs_id)
            return vs_id
        except Exception:
            pass
    vs = client.vector_stores.create(name=f"{project_name}_{pool}")
    pool_meta["vector_store_id"] = vs.id
    return vs.id


def ensure_assistant(meta: dict, pool: str, project_name: str) -> str:
    """Ensure an assistant exists for the pool, create if needed."""
    pool_meta = meta["pools"][pool]
    asst_id = pool_meta.get("assistant_id", "")
    vs_id = pool_meta.get("vector_store_id", "")
    if asst_id:
        try:
            client.beta.assistants.retrieve(asst_id)
            # Update vector store attachment in case it changed
            client.beta.assistants.update(
                assistant_id=asst_id,
                tool_resources={"file_search": {"vector_store_ids": [vs_id] if vs_id else []}},
            )
            return asst_id
        except Exception:
            pass
    pool_label = "Main Documents" if pool == "main" else "Clarification Documents"
    asst = client.beta.assistants.create(
        name=f"{project_name} — {pool_label}",
        instructions=SYSTEM_PROMPT,
        model=CHAT_MODEL,
        tools=[{"type": "file_search"}],
        tool_resources={"file_search": {"vector_store_ids": [vs_id] if vs_id else []}},
    )
    pool_meta["assistant_id"] = asst.id
    return asst.id


# ── Background task tracking ─────────────────────────────────────────────────

_tasks: dict[str, dict[str, Any]] = {}
_sync_lock = threading.Lock()


def _task_key(project_id: str, pool: str) -> str:
    return f"{project_id}_{pool}"


# ── Sync logic ───────────────────────────────────────────────────────────────

def sync_pool(project_id: str, pool: str, rebuild: bool = False):
    """Upload new/changed files to OpenAI vector store for a project pool.
    
    Sync logic:
    - Rebuild=True: deletes everything and re-indexes all files from scratch
    - Rebuild=False (Sync): only uploads files that are new or changed
      * For Excel files: also re-uploads if they were indexed with old code
        (detected by checking for 'excel_converted' flag in metadata)
    """
    key = _task_key(project_id, pool)
    try:
        _tasks[key] = {"status": "running", "progress": "Starting...", "result": None}
        project_dir = get_project_dir(project_id)
        meta = load_meta(project_dir)
        pool_meta = meta["pools"][pool]
        docs_dir = project_dir / f"documents_{pool}"
        docs_dir.mkdir(parents=True, exist_ok=True)

        # Ensure vector store
        _tasks[key]["progress"] = "Preparing vector store..."

        # Check if the existing vector store is still valid
        vs_id = pool_meta.get("vector_store_id", "")
        vs_is_valid = False
        if vs_id:
            try:
                client.vector_stores.retrieve(vs_id)
                vs_is_valid = True
            except Exception:
                log.warning(f"Vector store {vs_id} no longer exists — will recreate")
                vs_is_valid = False

        # If vector store is gone, treat as rebuild (all files need re-upload)
        if not vs_is_valid and not rebuild:
            log.info("Vector store missing — forcing full re-upload")
            pool_meta["vector_store_id"] = ""
            pool_meta["files"] = {}
            old_asst = pool_meta.get("assistant_id", "")
            if old_asst:
                try:
                    client.beta.assistants.delete(old_asst)
                except Exception:
                    pass
            pool_meta["assistant_id"] = ""
            save_meta(project_dir, meta)

        if rebuild:
            # Delete old vector store and create new one
            old_vs = pool_meta.get("vector_store_id", "")
            if old_vs:
                try:
                    client.vector_stores.delete(old_vs)
                except Exception:
                    pass
            pool_meta["vector_store_id"] = ""
            pool_meta["files"] = {}
            # Delete old assistant too
            old_asst = pool_meta.get("assistant_id", "")
            if old_asst:
                try:
                    client.beta.assistants.delete(old_asst)
                except Exception:
                    pass
            pool_meta["assistant_id"] = ""
            save_meta(project_dir, meta)

        vs_id = ensure_vector_store(meta, pool, project_id)
        save_meta(project_dir, meta)

        # Scan local files
        existing_files = pool_meta.get("files", {})
        local_files = {}
        for ext in ("*.pdf", "*.xlsx", "*.xls"):
            for p in docs_dir.glob(ext):
                local_files[p.name] = p

        log_entries = []
        uploaded_count = 0
        skipped_count = 0

        # Determine which files need uploading
        for fname, fpath in sorted(local_files.items()):
            needs_upload = False
            reason = ""

            if fname not in existing_files:
                # Brand new file — never indexed
                needs_upload = True
                reason = "new file"
            elif existing_files[fname].get("size") != fpath.stat().st_size:
                # File size changed on disk
                needs_upload = True
                reason = "file changed"
            elif fname.lower().endswith((".xlsx", ".xls")) and not existing_files[fname].get("excel_converted"):
                # Excel file was indexed with OLD code (before conversion fix)
                # It needs to be re-uploaded with proper text conversion
                needs_upload = True
                reason = "excel needs conversion"
                # Delete the old (broken) file from OpenAI
                old_fid = existing_files[fname].get("openai_file_id", "")
                if old_fid:
                    try:
                        client.files.delete(old_fid)
                    except Exception:
                        pass
            else:
                # File is in metadata with same size — already indexed
                skipped_count += 1
                continue

            _tasks[key]["progress"] = f"Uploading ({reason}): {fname}"
            log.info(f"Uploading {fname} to OpenAI ({reason})...")

            try:
                # For Excel files, convert to text first (OpenAI doesn't support .xlsx)
                if fname.lower().endswith((".xlsx", ".xls")):
                    log.info(f"Converting Excel file to text: {fname}")
                    _tasks[key]["progress"] = f"Converting Excel: {fname}"
                    text_content = excel_to_text(fpath)
                    if not text_content or not text_content.strip():
                        log_entries.append(f"Skipped (empty Excel): {fname}")
                        log.warning(f"Excel file {fname} produced empty text — skipping")
                        continue
                    
                    # Create a clean .txt filename: report.xlsx -> report_xlsx.txt
                    base, ext_part = fname.rsplit(".", 1)
                    clean_name = f"{base}_{ext_part}.txt"
                    log.info(f"Excel {fname} converted: {len(text_content)} chars -> uploading as {clean_name}")
                    
                    # Write to temp file with clean name
                    tmp_dir = tempfile.mkdtemp()
                    tmp_path = os.path.join(tmp_dir, clean_name)
                    with open(tmp_path, "w", encoding="utf-8") as f:
                        f.write(text_content)
                    
                    _tasks[key]["progress"] = f"Uploading converted: {fname}"
                    with open(tmp_path, "rb") as f:
                        oai_file = client.files.create(file=f, purpose="assistants")
                    
                    # Clean up temp file
                    try:
                        os.unlink(tmp_path)
                        os.rmdir(tmp_dir)
                    except Exception:
                        pass
                    
                    log.info(f"Excel {fname} uploaded to OpenAI as file_id={oai_file.id}")
                else:
                    # PDF files — upload directly
                    with open(fpath, "rb") as f:
                        oai_file = client.files.create(file=f, purpose="assistants")

                # Add to vector store
                _tasks[key]["progress"] = f"Indexing: {fname}"
                log.info(f"Adding {fname} to vector store {vs_id}...")
                vs_result = client.vector_stores.files.create_and_poll(
                    vector_store_id=vs_id, file_id=oai_file.id
                )
                
                # Check if indexing was successful
                if hasattr(vs_result, 'status') and vs_result.status == 'failed':
                    error_msg = getattr(vs_result, 'last_error', 'Unknown error')
                    log_entries.append(f"Index FAILED: {fname} — {error_msg}")
                    log.error(f"Vector store indexing failed for {fname}: {error_msg}")
                    # Do NOT save to metadata so it gets retried next sync
                    continue

                file_meta = {
                    "openai_file_id": oai_file.id,
                    "size": fpath.stat().st_size,
                }
                # Mark Excel files as properly converted so we don't re-upload next time
                if fname.lower().endswith((".xlsx", ".xls")):
                    file_meta["excel_converted"] = True
                
                existing_files[fname] = file_meta
                uploaded_count += 1
                file_type = "Excel->Text" if fname.lower().endswith((".xlsx", ".xls")) else "PDF"
                log_entries.append(f"Indexed ({file_type}): {fname}")
                log.info(f"Successfully indexed {fname} ({file_type})")

            except Exception as e:
                log_entries.append(f"Error uploading {fname}: {str(e)}")
                log.error(f"Error uploading {fname}: {e}", exc_info=True)
                # Remove from metadata so it gets retried next sync
                if fname in existing_files:
                    del existing_files[fname]

        # Remove files that no longer exist locally
        removed = []
        for fname in list(existing_files.keys()):
            if fname not in local_files:
                file_info = existing_files[fname]
                try:
                    client.files.delete(file_info["openai_file_id"])
                except Exception:
                    pass
                del existing_files[fname]
                removed.append(fname)
                log_entries.append(f"Removed: {fname}")

        pool_meta["files"] = existing_files
        save_meta(project_dir, meta)

        # Ensure assistant exists and is linked
        ensure_assistant(meta, pool, project_id)
        save_meta(project_dir, meta)

        result = {
            "files": len(existing_files),
            "uploaded": uploaded_count,
            "skipped": skipped_count,
            "removed": len(removed),
            "log": log_entries,
        }
        _tasks[key] = {"status": "done", "progress": "Complete", "result": result}
        log.info(f"Sync complete for {project_id}/{pool}: {len(existing_files)} total, {uploaded_count} new, {skipped_count} already indexed")

    except Exception as e:
        log.error(f"Sync error: {e}", exc_info=True)
        _tasks[key] = {"status": "error", "progress": str(e), "result": None}


# ── Chat logic ───────────────────────────────────────────────────────────────

# Thread cache: per project+pool, we keep a thread_id for conversation continuity
_threads: dict[str, str] = {}


def chat_with_pool(project_id: str, pool: str, user_message: str) -> dict:
    """Send a message to the assistant for a project pool and get a response."""
    project_dir = get_project_dir(project_id)
    meta = load_meta(project_dir)
    pool_meta = meta["pools"][pool]

    asst_id = pool_meta.get("assistant_id", "")
    if not asst_id:
        return {"answer": "No assistant found. Please sync documents first.", "sources": []}

    # Get or create thread
    thread_key = f"{project_id}_{pool}"
    thread_id = _threads.get(thread_key)
    if thread_id:
        try:
            client.beta.threads.retrieve(thread_id)
        except Exception:
            thread_id = None

    if not thread_id:
        thread = client.beta.threads.create()
        thread_id = thread.id
        _threads[thread_key] = thread_id

    # Add user message
    client.beta.threads.messages.create(
        thread_id=thread_id, role="user", content=user_message
    )

    # Run the assistant
    run = client.beta.threads.runs.create_and_poll(
        thread_id=thread_id,
        assistant_id=asst_id,
        timeout=300,
    )

    if run.status != "completed":
        return {
            "answer": f"Run failed with status: {run.status}. Please try again.",
            "sources": [],
        }

    # Get the assistant's response
    messages = client.beta.threads.messages.list(thread_id=thread_id, order="desc", limit=1)
    if not messages.data:
        return {"answer": "No response received.", "sources": []}

    msg = messages.data[0]
    answer_text = ""
    sources = []

    for block in msg.content:
        if block.type == "text":
            text_val = block.text.value
            # Process citations
            if block.text.annotations:
                for ann in block.text.annotations:
                    if hasattr(ann, "file_citation") and ann.file_citation:
                        try:
                            cited_file = client.files.retrieve(ann.file_citation.file_id)
                            source_name = cited_file.filename
                            # Clean up the source name
                            # Convert back: report_xlsx.txt → report.xlsx
                            if source_name.endswith("_xlsx.txt"):
                                source_name = source_name[:-9] + ".xlsx"
                            elif source_name.endswith("_xls.txt"):
                                source_name = source_name[:-8] + ".xls"
                            # Remove any temp file prefixes
                            if source_name.startswith("_"):
                                source_name = source_name.lstrip("_")
                            if source_name not in sources:
                                sources.append(source_name)
                        except Exception:
                            pass
                    text_val = text_val.replace(ann.text, "")
            answer_text += text_val

    return {"answer": answer_text.strip(), "sources": sources}


# ── Flask App ────────────────────────────────────────────────────────────────

app = Flask(__name__)
app.secret_key = SECRET_KEY


# ── Password protection middleware ───────────────────────────────────────────

@app.before_request
def check_auth():
    if not APP_PASSWORD:
        return None
    if request.path.startswith("/static") or request.path == "/login":
        return None
    if session.get("authenticated"):
        return None
    if request.method == "POST" and request.path == "/login":
        return None
    return redirect(url_for("login_page"))


@app.route("/login", methods=["GET", "POST"])
def login_page():
    if not APP_PASSWORD:
        return redirect("/")
    error = ""
    if request.method == "POST":
        pw = request.form.get("password", "")
        if pw == APP_PASSWORD:
            session["authenticated"] = True
            return redirect("/")
        error = "Incorrect password"
    return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Login — Project RAG Pro</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:#0f172a;color:#e2e8f0;font-family:system-ui,-apple-system,sans-serif;display:flex;align-items:center;justify-content:center;min-height:100vh}}
.card{{background:#1e293b;border-radius:12px;padding:2.5rem;width:360px;box-shadow:0 8px 32px rgba(0,0,0,.4)}}
h1{{font-size:1.4rem;margin-bottom:.5rem;text-align:center}}
.sub{{color:#94a3b8;text-align:center;margin-bottom:1.5rem;font-size:.9rem}}
input{{width:100%;padding:.75rem 1rem;border-radius:8px;border:1px solid #334155;background:#0f172a;color:#e2e8f0;font-size:1rem;margin-bottom:1rem}}
button{{width:100%;padding:.75rem;border-radius:8px;border:none;background:#3b82f6;color:#fff;font-size:1rem;cursor:pointer;font-weight:600}}
button:hover{{background:#2563eb}}
.err{{color:#f87171;text-align:center;margin-bottom:1rem;font-size:.85rem}}
</style></head><body>
<div class="card"><h1>Project RAG Pro</h1><p class="sub">Enter password to continue</p>
{'<p class="err">'+error+'</p>' if error else ''}
<form method="POST"><input type="password" name="password" placeholder="Password" autofocus>
<button type="submit">Login</button></form></div></body></html>"""


# ── API Routes ───────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


# ── Projects CRUD ────────────────────────────────────────────────────────────

@app.route("/api/projects", methods=["GET"])
def list_projects():
    BASE_DIR.mkdir(parents=True, exist_ok=True)
    projects = []
    for d in sorted(BASE_DIR.iterdir()):
        if d.is_dir() and (d / "metadata.json").exists():
            meta = load_meta(d)
            main_files = len(meta["pools"]["main"].get("files", {}))
            clar_files = len(meta["pools"]["clarification"].get("files", {}))
            projects.append({
                "id": d.name,
                "description": meta.get("description", ""),
                "created": meta.get("created", ""),
                "main_files": main_files,
                "clar_files": clar_files,
            })
    return jsonify(projects)


@app.route("/api/projects", methods=["POST"])
def create_project():
    data = request.get_json(force=True)
    name = data.get("name", "").strip().replace(" ", "_")
    if not name:
        return jsonify({"error": "Name required"}), 400
    project_dir = get_project_dir(name)
    if project_dir.exists():
        return jsonify({"error": "Project already exists"}), 409
    project_dir.mkdir(parents=True, exist_ok=True)
    (project_dir / "documents_main").mkdir(exist_ok=True)
    (project_dir / "documents_clarification").mkdir(exist_ok=True)
    meta = load_meta(project_dir)
    meta["description"] = data.get("description", "")
    meta["created"] = time.strftime("%Y-%m-%d")
    save_meta(project_dir, meta)
    return jsonify({"id": name})


@app.route("/api/projects/<project_id>", methods=["DELETE"])
def delete_project(project_id):
    project_dir = get_project_dir(project_id)
    if not project_dir.exists():
        return jsonify({"error": "Not found"}), 404
    # Clean up OpenAI resources
    meta = load_meta(project_dir)
    for pool in POOL_TYPES:
        pool_meta = meta["pools"].get(pool, {})
        # Delete assistant
        asst_id = pool_meta.get("assistant_id", "")
        if asst_id:
            try:
                client.beta.assistants.delete(asst_id)
            except Exception:
                pass
        # Delete vector store
        vs_id = pool_meta.get("vector_store_id", "")
        if vs_id:
            try:
                client.vector_stores.delete(vs_id)
            except Exception:
                pass
        # Delete files
        for finfo in pool_meta.get("files", {}).values():
            try:
                client.files.delete(finfo["openai_file_id"])
            except Exception:
                pass
    shutil.rmtree(project_dir)
    return jsonify({"deleted": True})


# ── Project status ───────────────────────────────────────────────────────────

@app.route("/api/projects/<project_id>/status", methods=["GET"])
def project_status(project_id):
    project_dir = get_project_dir(project_id)
    if not project_dir.exists():
        return jsonify({"error": "Not found"}), 404
    meta = load_meta(project_dir)
    result = {}
    for pool in POOL_TYPES:
        pool_meta = meta["pools"].get(pool, {})
        indexed_files = pool_meta.get("files", {})
        docs_dir = project_dir / f"documents_{pool}"
        local_count = 0
        if docs_dir.exists():
            for ext in ("*.pdf", "*.xlsx", "*.xls"):
                local_count += len(list(docs_dir.glob(ext)))
        result[pool] = {
            "indexed_files": len(indexed_files),
            "local_files": local_count,
            "has_assistant": bool(pool_meta.get("assistant_id")),
            "files": [
                {"name": fname, "indexed": fname in indexed_files}
                for fname in sorted(
                    set(
                        list(indexed_files.keys())
                        + ([p.name for ext in ("*.pdf", "*.xlsx", "*.xls") for p in docs_dir.glob(ext)] if docs_dir.exists() else [])
                    )
                )
            ],
        }
    return jsonify(result)


# ── File upload ──────────────────────────────────────────────────────────────

@app.route("/api/projects/<project_id>/<pool>/upload", methods=["POST"])
def upload_file(project_id, pool):
    if pool not in POOL_TYPES:
        return jsonify({"error": f"Invalid pool: {pool}"}), 400
    project_dir = get_project_dir(project_id)
    if not project_dir.exists():
        return jsonify({"error": "Project not found"}), 404
    docs_dir = project_dir / f"documents_{pool}"
    docs_dir.mkdir(parents=True, exist_ok=True)
    saved = []
    for f in request.files.getlist("files"):
        fname = f.filename or "unnamed"
        dest = docs_dir / fname
        f.save(str(dest))
        saved.append(fname)
    return jsonify({"uploaded": saved})


# ── File delete ──────────────────────────────────────────────────────────────

@app.route("/api/projects/<project_id>/<pool>/files/<filename>", methods=["DELETE"])
def delete_file(project_id, pool, filename):
    if pool not in POOL_TYPES:
        return jsonify({"error": f"Invalid pool: {pool}"}), 400
    project_dir = get_project_dir(project_id)
    docs_dir = project_dir / f"documents_{pool}"
    fpath = docs_dir / filename
    if fpath.exists():
        fpath.unlink()
    # Also remove from OpenAI
    meta = load_meta(project_dir)
    pool_meta = meta["pools"][pool]
    files_dict = pool_meta.get("files", {})
    if filename in files_dict:
        try:
            client.files.delete(files_dict[filename]["openai_file_id"])
        except Exception:
            pass
        del files_dict[filename]
        save_meta(project_dir, meta)
    return jsonify({"deleted": filename})


# ── Sync / Rebuild (background) ──────────────────────────────────────────────

@app.route("/api/projects/<project_id>/<pool>/sync", methods=["POST"])
def sync_route(project_id, pool):
    if pool not in POOL_TYPES:
        return jsonify({"error": f"Invalid pool: {pool}"}), 400
    project_dir = get_project_dir(project_id)
    if not project_dir.exists():
        return jsonify({"error": "Project not found"}), 404
    key = _task_key(project_id, pool)
    existing = _tasks.get(key, {})
    if existing.get("status") == "running":
        return jsonify({"error": "Sync already in progress"}), 409
    t = threading.Thread(target=sync_pool, args=(project_id, pool, False), daemon=True)
    t.start()
    return jsonify({"started": True, "task_key": key})


@app.route("/api/projects/<project_id>/<pool>/rebuild", methods=["POST"])
def rebuild_route(project_id, pool):
    if pool not in POOL_TYPES:
        return jsonify({"error": f"Invalid pool: {pool}"}), 400
    project_dir = get_project_dir(project_id)
    if not project_dir.exists():
        return jsonify({"error": "Project not found"}), 404
    key = _task_key(project_id, pool)
    existing = _tasks.get(key, {})
    if existing.get("status") == "running":
        return jsonify({"error": "Sync already in progress"}), 409
    t = threading.Thread(target=sync_pool, args=(project_id, pool, True), daemon=True)
    t.start()
    return jsonify({"started": True, "task_key": key})


@app.route("/api/projects/<project_id>/<pool>/task_status", methods=["GET"])
def task_status_route(project_id, pool):
    key = _task_key(project_id, pool)
    task = _tasks.get(key, {"status": "idle", "progress": "", "result": None})
    return jsonify(task)


# ── Chat ─────────────────────────────────────────────────────────────────────

@app.route("/api/projects/<project_id>/<pool>/chat", methods=["POST"])
def chat_route(project_id, pool):
    if pool not in POOL_TYPES:
        return jsonify({"error": f"Invalid pool: {pool}"}), 400
    project_dir = get_project_dir(project_id)
    if not project_dir.exists():
        return jsonify({"error": "Project not found"}), 404
    data = request.get_json(force=True)
    question = data.get("question", "").strip()
    if not question:
        return jsonify({"error": "Empty question"}), 400
    result = chat_with_pool(project_id, pool, question)
    return jsonify(result)


# ── Reset chat thread ────────────────────────────────────────────────────────

@app.route("/api/projects/<project_id>/<pool>/reset_chat", methods=["POST"])
def reset_chat_route(project_id, pool):
    thread_key = f"{project_id}_{pool}"
    old_id = _threads.pop(thread_key, None)
    if old_id:
        try:
            client.beta.threads.delete(old_id)
        except Exception:
            pass
    return jsonify({"reset": True})


# ── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    BASE_DIR.mkdir(parents=True, exist_ok=True)
    print(f"\n  Open http://localhost:8080 in your browser.\n")
    app.run(host="127.0.0.1", port=8080, debug=False)
